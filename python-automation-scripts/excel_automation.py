from openpyxl import Workbook, load_workbook
from pathlib import Path

def create_sample(path="sample.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Score"])
    ws.append(["Alice", 90])
    ws.append(["Bob", 85])
    wb.save(path)
    return path

def read_scores(path):
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.values)
    return rows[1:]

if __name__ == "__main__":
    p = create_sample()
    print("Created:", p)
    print(read_scores(p))